import time
import requests
from bs4 import BeautifulSoup
import urllib.parse
import re
import math
import io
import pandas as pd
from datetime import datetime, timedelta
from concurrent.futures import ThreadPoolExecutor, as_completed
import streamlit as st

# --- ページの設定（タイトルやアイコン） ---
st.set_page_config(page_title="minne市場リサーチツール", page_icon="🛍️", layout="wide")

# 🎨 画面をギュッと引き締めるコンパクトデザイン ＆ 文字色カスタム設定
st.markdown("""
    <style>
    html, body, [class*="css"] { font-size: 14px !important; }
    .block-container { padding-top: 1.5rem !important; padding-bottom: 1rem !important; }
    h1 { font-size: 24px !important; font-weight: 700 !important; margin-bottom: 5px !important; }
    div[data-testid="stVerticalBlock"] > div { padding-bottom: 4px !important; }
    .stTextInput input, .stNumberInput input, .stDateInput input { padding: 6px 10px !important; font-size: 13px !important; }
    .stMarkdown p { margin-bottom: 2px !important; }
    
    /* サイドバーの独自テキストを黒くするための設定 */
    .custom-sidebar-label {
        color: #111111 !important;
        font-weight: 600 !important;
        font-size: 13px !important;
        margin-top: 10px !important;
        margin-bottom: 5px !important;
        display: block;
    }
    /* カレンダー入力欄の隙間をさらに詰めて見やすくする設定 */
    div[data-testid="stDateInput"] label { display: none !important; }
    </style>
""", unsafe_allow_html=True)

st.title("🛍️ minne市場リサーチツール")
st.caption("キーワード、またはminneの検索結果URLから売れ行きやライバル作品を爆速で一括解析します。")
st.write("---")

# 📲 LINE公式アカウントからあなたへ通知を飛ばす関数
def send_line_notification(search_target, limit_count):
    LINE_ACCESS_TOKEN = "SsJj64qF912H/fusrwNgsiMS6bgJqv5C9i5Rx1HlHAmux8AmFlC7Q9Pnx5pbQD/4LXbi2ftiFf1zalCCDcGQAcXBxfakpnkBPLZkKzn5G2gbuQc2vkcn2GbCJ2Yf1HmfEWQoo8KbqqJn4/tsoPr4TwdB04t89/1O/w1cDnyilFU="
    LINE_USER_ID = "Ub5228833332f8fd37bbd3d9072853f2c"
    
    url = "https://api.line.me/v2/bot/message/push"
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {LINE_ACCESS_TOKEN}"
    }
    
    message_text = (
        f"🛍️ 【minneツール】利用通知\n\n"
        f"今、誰かがリサーチを開始したよ！\n"
        f"---------------------\n"
        f"▼ 検索内容:\n{search_target}\n\n"
        f"▼ 解析上限: {limit_count} 件"
    )
    
    payload = {"to": LINE_USER_ID, "messages": [{"type": "text", "text": message_text}]}
    try: requests.post(url, headers=headers, json=payload, timeout=5)
    except: pass

# --- 詳細データ抽出関数 ---
def get_minne_perfect_details(product_url):
    headers = {"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"}
    max_retries = 3
    for attempt in range(1, max_retries + 1):
        try:
            time.sleep(0.2)
            res = requests.get(product_url, headers=headers, timeout=10)
            if res.status_code != 200: continue
            soup = BeautifulSoup(res.text, "html.parser")
            price_tag = soup.find(class_=lambda x: x and x.startswith("MinneProductSummary_price__"))
            if not price_tag: continue
            price = price_tag.text.strip()
            if price and "円" not in price: price = f"{price}円"
            break
        except:
            if attempt == max_retries: return {}
            continue
    else: return {}

    try:
        shop_tag = soup.find(class_=lambda x: x and x.startswith("MinneProductSummary_shop-name__"))
        shop_name = shop_tag.text.strip() if shop_tag else "取得失敗"
        chip_tags = soup.find_all(class_=lambda x: x and x.startswith("MyChip_chip-gray__"))
        tags = [tag.text.strip() for tag in chip_tags if tag.text.strip().startswith("#")]
        hashtag_str = ", ".join(tags) if tags else "なし"
        
        favorite_count = "0件"
        fav_tag = soup.find(class_=lambda x: x and x.startswith("MinneProductSummary_favorite-count__"))
        if fav_tag:
            fav_text = fav_tag.text.strip()
            clean_fav_text = fav_text.replace(',', '')
            if clean_fav_text.isdigit():
                favorite_count = f"{clean_fav_text}件"

        related_count, shop_review_count, related_num, shop_review_num = "0件", "0件", 0, 0
        sidebar = soup.find(class_=lambda x: x and (x.startswith("ProductSinglePage-sidebar-shop-wrapper__") or x.startswith("ProductSinglePage_sidebar-shop-wrapper__")))
        if sidebar:
            sidebar_text = sidebar.text.strip()
            related_match = re.search(r'関連レビュー\s*(\d+)', sidebar_text)
            shop_match = re.search(r'ショップレビュー\s*(\d+)', sidebar_text)
            if related_match:
                related_num = int(related_match.group(1))
                related_count = f"{related_num}件"
            if shop_match:
                shop_review_num = int(shop_match.group(1))
                shop_review_count = f"{shop_review_num}件"

        date_list = []
        if related_num == 0:
            date_list = ["なし", "なし", "なし"]
        else:
            review_dates = soup.find_all(class_=lambda x: x and x.startswith("MinneReviewCard_reviewDate__"))
            for i in range(3):
                if i < len(review_dates): date_list.append(review_dates[i].text.strip())
                else: date_list.append("なし")
        
        first_shop_review_date = "なし"
        if shop_review_num > 0 and shop_tag and shop_tag.get("href"):
            try:
                raw_path = shop_tag.get("href").split('?')[0].strip('/')
                shop_id = raw_path.split('/')[-1] 
                calculated_last_page = math.ceil(shop_review_num / 10)
                
                for retry_offset in range(10):
                    target_page = calculated_last_page - retry_offset
                    if target_page <= 0: break
                    reviews_url = f"https://minne.com/{shop_id}/reviews?page={target_page}"
                    
                    time.sleep(0.1)
                    rev_res = requests.get(reviews_url, headers=headers, timeout=10)
                    if rev_res.status_code == 200:
                        found_dates = re.findall(r'\d{4}/\d{2}/\d{2}', rev_res.text)
                        if found_dates:
                            first_shop_review_date = min(found_dates)
                            break
                        else:
                            first_shop_review_date = "レビュー日なし"
                    else:
                        first_shop_review_date = f"エラー({rev_res.status_code})"
                        continue
            except:
                first_shop_review_date = "解析失敗"
        
        return {
            "ショップ名": shop_name, 
            "価格": price, 
            "ハッシュタグ": hashtag_str, 
            "お気に入り数": favorite_count,
            "関連レビュー数": related_count, 
            "ショップレビュー数": shop_review_count, 
            "最初のショップレビュー日": first_shop_review_date, 
            "レビュー日1": date_list[0], 
            "レビュー日2": date_list[1], 
            "レビュー日3": date_list[2]
        }
    except: return {}

# --- セッション状態の初期化 ---
if "df_scraped_raw" not in st.session_state:
    st.session_state.df_scraped_raw = None

# --- 🛰️ 画面（サイドバー）の設定 ---
st.sidebar.header("🔍 1. データ取得元の指定")
target_input = st.sidebar.text_input("キーワード または 検索結果URL", value="")
limit = st.sidebar.number_input("解析する件数上限", min_value=10, max_value=500, value=40, step=10)

# 🚀 リサーチ開始ボタン（このボタンはデータ取得のみを行う）
if st.sidebar.button("リサーチを開始する", type="primary", use_container_width=True):
    search_log_text = target_input if target_input.strip() != "" else "（未入力・空欄リサーチ）"
    send_line_notification(search_log_text, limit)

    headers = {"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"}
    if target_input.startswith("http://") or target_input.startswith("https://"):
        st.info("🔗 URLモードで実行中...")
        base_url = target_input
    else:
        st.info(f"🔍 キーワードモード（{target_input}）で実行中...")
        encoded_keyword = urllib.parse.quote(target_input)
        base_url = f"https://minne.com/category/saleonly?input_method=typing&q={encoded_keyword}&commit=%E6%A4%9C%E7%B4%A2&search_type=normal"
    
    product_urls_with_titles = []
    page = 1
    seen_urls = set()
    status_text = st.empty()
    
    while len(product_urls_with_titles) < limit:
        search_url = f"{base_url}&page={page}" if "?" in base_url else f"{base_url}?page={page}"
        status_text.text(f"📄 minneの {page} ページ目を読み込み中... (現在 {len(product_urls_with_titles)} 件の候補)")
        try:
            response = requests.get(search_url, headers=headers)
            if response.status_code != 200: break
            soup = BeautifulSoup(response.text, "html.parser")
            products = soup.find_all("a", class_=lambda x: x and (x.startswith("MinneProductCard_grid__") or x.startswith("ProductGrid_item__") or "product-card" in x.lower()))
            if not products: products = soup.find_all('a', href=re.compile(r'/items/\d+'))
            if not products: break
            
            page_added = 0
            for product in products:
                href = product.get("href")
                if not href: continue
                p_url = "https://minne.com" + href.split('?')[0] if href.startswith("/items/") else (href.split('?')[0] if "minne.com/items/" in href else None)
                if p_url and p_url not in seen_urls:
                    seen_urls.add(p_url)
                    title_tag = product.find(class_=lambda x: x and "title" in x.lower())
                    title = title_tag.text.strip() if title_tag else "商品名（個別解析で取得）"
                    product_urls_with_titles.append((title, p_url))
                    page_added += 1
            if page_added == 0: break
            page += 1
            time.sleep(0.3)
        except:
            break

    total_to_scrape = min(limit, len(product_urls_with_titles))
    if total_to_scrape == 0:
        st.error("❌ 条件に合う商品が1件も取得できませんでした。")
        st.session_state.df_scraped_raw = None
    else:
        status_text.text(f"📦 候補 {total_to_scrape} 件の詳細データを並行解析中...（しばらくお待ちください）")
        progress_bar = st.progress(0)
        raw_results = [None] * total_to_scrape
        completed_count = 0
        
        with ThreadPoolExecutor(max_workers=5) as executor:
            future_to_info = {executor.submit(get_minne_perfect_details, url): (idx, title, url) for idx, (title, url) in enumerate(product_urls_with_titles[:total_to_scrape])}
            for future in as_completed(future_to_info):
                idx, title, url = future_to_info[future]
                try: details = future.result()
                except: details = {}
                
                display_title = "作品（詳細はURLへ）" if title == "商品名（個別解析で取得）" and "ショップ名" in details else title
                raw_results[idx] = {
                    "ショップ名": details.get("ショップ名", "取得失敗"), "商品名": display_title, "価格": details.get("価格", "価格なし"),
                    "URL": url, 
                    "お気に入り数": details.get("お気に入り数", "0件"),
                    "関連レビュー数": details.get("関連レビュー数", "0件"),
                    "最新の関連レビュー日": details.get("レビュー日1", "なし"), "2件目の関連レビュー日": details.get("レビュー日2", "なし"), "3件目の関連レビュー日": details.get("レビュー日3", "なし"),
                    "ショップレビュー数": details.get("ショップレビュー数", "0件"), 
                    "最初のショップレビュー日": details.get("最初のショップレビュー日", "なし"), 
                    "ハッシュタグ": details.get("ハッシュタグ", "なし")
                }
                completed_count += 1
                progress_bar.progress(completed_count / total_to_scrape)
        
        status_text.empty()
        progress_bar.empty()
        st.session_state.df_scraped_raw = pd.DataFrame(raw_results)

st.sidebar.write("---")
# 🛠️ タイトルを修正
st.sidebar.header("⏳ 2. 抽出結果の絞り込み")
st.sidebar.caption("※抽出後に各数値を変更すると、自動で表が絞り込まれます。")

# 🛠️ 価格フィルターを横並び（2列）に修正
st.sidebar.markdown('<span class="custom-sidebar-label">💰 価格帯 (円)</span>', unsafe_allow_html=True)
col_p1, col_p2 = st.sidebar.columns(2)
with col_p1: min_p = st.number_input("最低", min_value=0, value=0, step=100, key="min_p")
with col_p2: max_p = st.number_input("最高", min_value=0, value=100000, step=100, key="max_p")

st.sidebar.markdown('<span class="custom-sidebar-label">❤️ お気に入り数</span>', unsafe_allow_html=True)
col_fav1, col_fav2 = st.sidebar.columns(2)
with col_fav1: min_fav = st.number_input("最低", min_value=0, value=0, key="min_fav")
with col_fav2: max_fav = st.number_input("最高", min_value=0, value=99999, key="max_fav")

st.sidebar.markdown('<span class="custom-sidebar-label">📊 関連レビュー数</span>', unsafe_allow_html=True)
col1, col2 = st.sidebar.columns(2)
with col1: min_rev = st.number_input("最低", min_value=0, value=0, key="min_rev")
with col2: max_rev = st.number_input("最高", min_value=0, value=9999, key="max_rev")

today = datetime.now()
seven_days_ago = today - timedelta(days=7)

use_date_filter_1 = st.sidebar.checkbox("最新の関連レビュー日を指定する", value=False)
date_range_1 = st.sidebar.date_input("最新レビュー範囲", value=(seven_days_ago, today), max_value=today, key="dr_related_1") if use_date_filter_1 else None

use_date_filter_2 = st.sidebar.checkbox("2件目の関連レビュー日を指定する", value=False)
date_range_2 = st.sidebar.date_input("2件目レビュー範囲", value=(seven_days_ago, today), max_value=today, key="dr_related_2") if use_date_filter_2 else None

use_date_filter_3 = st.sidebar.checkbox("3件目の関連レビュー日を指定する", value=False)
date_range_3 = st.sidebar.date_input("3件目レビュー範囲", value=(seven_days_ago, today), max_value=today, key="dr_related_3") if use_date_filter_3 else None

st.sidebar.markdown('<span class="custom-sidebar-label">🏪 ショップレビュー数</span>', unsafe_allow_html=True)
col3, col4 = st.sidebar.columns(2)
with col3: min_shop_rev = st.number_input("最低", min_value=0, value=0, key="min_shop_rev")
with col4: max_shop_rev = st.number_input("最高", min_value=0, value=99999, key="max_shop_rev")

use_shop_date_filter = st.sidebar.checkbox("最初のショップレビュー日を指定する", value=False)
shop_date_range = st.sidebar.date_input("ショップ開始範囲", value=(today - timedelta(days=3652), today), max_value=today, key="date_range_shop") if use_shop_date_filter else None


# --- 📊 リアルタイムデータ表示処理 ---
if st.session_state.df_scraped_raw is not None:
    if use_date_filter_1 and (not date_range_1 or len(date_range_1) != 2): st.warning("⚠️ 最新の関連レビュー日の「開始日と終了日」を両方選択してください。"); st.stop()
    if use_date_filter_2 and (not date_range_2 or len(date_range_2) != 2): st.warning("⚠️ 2件目の関連レビュー日の「開始日と終了日」を両方選択してください。"); st.stop()
    if use_date_filter_3 and (not date_range_3 or len(date_range_3) != 2): st.warning("⚠️ 3件目の関連レビュー日の「開始日と終了日」を両方選択してください。"); st.stop()
    if use_shop_date_filter and (not shop_date_range or len(shop_date_range) != 2): st.warning("⚠️ 最初のショップレビュー日の「開始日と終了日」を両方選択してください。"); st.stop()

    df_filter = st.session_state.df_scraped_raw.copy()
    
    df_filter['価格_数値'] = pd.to_numeric(df_filter['価格'].str.replace('円', '').str.replace(',', '').str.strip(), errors='coerce').fillna(0).astype(int)
    df_filter['お気に入り_数値'] = pd.to_numeric(df_filter['お気に入り数'].str.replace('件', '').str.strip(), errors='coerce').fillna(0).astype(int)
    df_filter['関連レビュー_数値'] = pd.to_numeric(df_filter['関連レビュー数'].str.replace('件', '').str.strip(), errors='coerce').fillna(0).astype(int)
    df_filter['ショップレビュー_数値'] = pd.to_numeric(df_filter['ショップレビュー数'].str.replace('件', '').str.strip(), errors='coerce').fillna(0).astype(int)
    
    def clean_japanese_date(date_str):
        if not date_str or pd.isna(date_str) or date_str in ['なし', 'レビューなし', '解析失敗', 'レビュー日なし']: return pd.NaT
        try: return pd.to_datetime(date_str.replace('年', '/').replace('月', '/').replace('日', '').strip(), format='%Y/%m/%d')
        except: return pd.NaT
        
    df_filter['最新の関連レビュー日_日付'] = df_filter['最新の関連レビュー日'].apply(clean_japanese_date)
    df_filter['2件目の関連レビュー日_日付'] = df_filter['2件目の関連レビュー日'].apply(clean_japanese_date)
    df_filter['3件目の関連レビュー日_日付'] = df_filter['3件目の関連レビュー日'].apply(clean_japanese_date)
    df_filter['最初のショップレビュー日_日付'] = df_filter['最初のショップレビュー日'].apply(clean_japanese_date)
    
    query_condition = (
        (df_filter['価格_数値'] >= min_p) & (df_filter['価格_数値'] <= max_p) &
        (df_filter['お気に入り_数値'] >= min_fav) & (df_filter['お気に入り_数値'] <= max_fav) &
        (df_filter['関連レビュー_数値'] >= min_rev) & (df_filter['関連レビュー_数値'] <= max_rev) &
        (df_filter['ショップレビュー_数値'] >= min_shop_rev) & (df_filter['ショップレビュー_数値'] <= max_shop_rev)
    )
    
    if use_date_filter_1 and date_range_1:
        start_dt_1 = datetime.combine(date_range_1[0], datetime.min.time())
        end_dt_1 = datetime.combine(date_range_1[1], datetime.max.time())
        query_condition = query_condition & (df_filter['関連レビュー_数値'] >= 1) & (df_filter['最新の関連レビュー日_日付'] >= start_dt_1) & (df_filter['最新の関連レビュー日_日付'] <= end_dt_1)
        
    if use_date_filter_2 and date_range_2:
        start_dt_2 = datetime.combine(date_range_2[0], datetime.min.time())
        end_dt_2 = datetime.combine(date_range_2[1], datetime.max.time())
        query_condition = query_condition & (df_filter['関連レビュー_数値'] >= 2) & (df_filter['2件目の関連レビュー日_日付'] >= start_dt_2) & (df_filter['2件目の関連レビュー日_日付'] <= start_dt_2)
        
    if use_date_filter_3 and date_range_3:
        start_dt_3 = datetime.combine(date_range_3[0], datetime.min.time())
        end_dt_3 = datetime.combine(date_range_3[1], datetime.max.time())
        query_condition = query_condition & (df_filter['関連レビュー_数値'] >= 3) & (df_filter['3件目の関連レビュー日_日付'] >= start_dt_3) & (df_filter['3件目の関連レビュー日_日付'] <= end_dt_3)
        
    if use_shop_date_filter and shop_date_range:
        shop_start_dt = datetime.combine(shop_date_range[0], datetime.min.time())
        shop_end_dt = datetime.combine(shop_date_range[1], datetime.max.time())
        query_condition = query_condition & (df_filter['ショップレビュー_数値'] > 0) & (df_filter['最初のショップレビュー日_日付'] >= shop_start_dt) & (df_filter['最初のショップレビュー日_日付'] <= shop_end_dt)
        
    df_result = df_filter[query_condition]
    display_cols = ["ショップ名", "商品名", "価格", "URL", "お気に入り数", "関連レビュー数", "最新の関連レビュー日", "2件目の関連レビュー日", "3件目の関連レビュー日", "ショップレビュー数", "最初のショップレビュー日", "ハッシュタグ"]
    df_final = df_result[display_cols].copy()
    
    st.success(f"✨ フィルター適用中: 全 {len(st.session_state.df_scraped_raw)} 件中 {len(df_final)} 件を表示しています。")
    
    # --- Excel生成 ---
    output_excel = io.BytesIO()
    with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
        df_final.to_excel(writer, index=False, sheet_name='リサーチ結果')
        workbook  = writer.book
        worksheet = writer.sheets['リサーチ結果']
        max_row = len(df_final) + 1
        max_col = len(display_cols)
        worksheet.auto_filter.ref = f"A1:{chr(64 + max_col)}{max_row}"
        
        from openpyxl.styles import PatternFill, Font, Alignment
        header_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
        header_font = Font(name='Meiryo UI', size=11, bold=True, color='000000')
        url_font = Font(name='Meiryo UI', size=11, color='0563C1', underline='single')
        normal_font = Font(name='Meiryo UI', size=11, color='000000')
        
        for col_idx in range(1, max_col + 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
        for row in range(2, max_row + 1):
            for col in range(1, max_col + 1):
                cell = worksheet.cell(row=row, column=col)
                if col == 4: cell.font = url_font
                else: cell.font = normal_font
                    
        for col in worksheet.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                val_str = str(cell.value or '')
                cell_len = sum([(2 if ord(c) > 256 else 1) for c in val_str])
                if cell_len > max_len: max_len = cell_len
            worksheet.column_dimensions[col_letter].width = min(max(max_len + 5, 11), 32)
            
    processed_excel = output_excel.getvalue()
    
    st.download_button(
        label="📥 絞り込んだ結果をExcelでDL", 
        data=processed_excel, 
        file_name=f"minne_filtered_{datetime.now().strftime('%Y%m%d')}.xlsx", 
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    
    st.dataframe(
        df_final, 
        column_config={
            "商品名": st.column_config.TextColumn("商品名", width=250),  
            "URL": st.column_config.LinkColumn("URL", width=120),       
            "ショップ名": st.column_config.TextColumn("ショップ名", width=160),
            "価格": st.column_config.TextColumn("価格", width=110),
            "お気に入り数": st.column_config.TextColumn("お気に入り数", width=120),
            "関連レビュー数": st.column_config.TextColumn("関連レビュー数", width=130),
            "最新の関連レビュー日": st.column_config.TextColumn("最新の関連レビュー日", width=150),
            "2件目の関連レビュー日": st.column_config.TextColumn("2件目の関連レビュー日", width=150),
            "3件目の関連レビュー日": st.column_config.TextColumn("3件目の関連レビュー日", width=150),
            "ショップレビュー数": st.column_config.TextColumn("ショップレビュー数", width=150),
            "最初のショップレビュー日": st.column_config.TextColumn("最初のショップレビュー日", width=180), 
            "ハッシュタグ": st.column_config.TextColumn("ハッシュタグ", width=250),
        }, 
        use_container_width=True
    )
else:
    st.info("💡 左メニューの「1. データ取得元の指定」を入力し、「リサーチを開始する」ボタンを押すと解析が始まります。")
